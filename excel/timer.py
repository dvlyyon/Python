#!/usr/bin/python3

import sys
from openpyxl import Workbook
import datetime

from openpyxl.chart import (LineChart, BarChart, PieChart, Reference,)
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import DataPoint

def drawLineChart(ws, srow, erow, scol, ecol, position, onlyone=False, title="Time interval", ytitle="value (ms)", xtitle="Timestamp"):
    data1= Reference(ws,min_col=scol, min_row=srow, max_col=ecol, max_row=erow)

    c1 = LineChart()
    c1.title = f"{title} ({str(srow)} - {str(erow)})"
    c1.style = 13
    c1.y_axis.majorGridlines = None
    c1.y_axis.title = ytitle
    c1.x_axis.title = xtitle
    
    c1.add_data(data1, titles_from_data=True)
    
    s1 = c1.series[0]
    s1.graphicalProperties.line.solidFill = "FF0000"
    s1.graphicalProperties.line.width = 10000
    s1.smooth = True
    
    if not onlyone:
        s2 = c1.series[1]
        s2.graphicalProperties.line.solidFill = "0000FF"
        s2.graphicalProperties.line.width = 10000
        s2.smooth = True
    
    ws.add_chart(c1, position)

def drawBarChart(ws, srow, erow, scol, ecol, position, title="Interval Statistics Bar Chart", ytitle="Number", xtitle="Sampling interval range"):

    c1 = BarChart()
    c1.title = title
    c1.type = 'col'
#    c1.style = 10
    c1.y_axis.majorGridlines = None
    c1.y_axis.title = ytitle
    c1.x_axis.title = xtitle
    c1.legend = None
    
    data1= Reference(ws,min_col=scol, min_row=srow, max_col=ecol, max_row=erow)
    cats = Reference(ws,min_col=1,min_row=srow,max_row=erow)
    c1.add_data(data1)
    c1.set_categories(cats)
    #c1.shape=4
    ws.add_chart(c1, position)

def drawPieChart(ws, srow, erow, scol, ecol, position, title="Sampling time difference (%)"):

    c1 = PieChart()
    
    data1 = Reference(ws,min_col=scol, min_row=srow, max_col=ecol, max_row=erow)
    labels = Reference(ws,min_col=1,min_row=srow+1,max_row=erow)
    c1.add_data(data1, titles_from_data=True)
    c1.set_categories(labels)
    c1.title = title

    slice = DataPoint(idx=0, explosion=20)
    c1.series[0].data_points = [slice]

    ws.add_chart(c1, position)

wb = Workbook()
ws = wb.active
ws.title = "Timestamp Statistics"
file=open(sys.argv[1])
ws1 = wb.create_sheet("statistics")
ws2 = wb.create_sheet("statistics %")
ws3 = wb.create_sheet("statistics1 %")

sampleInterval=int(sys.argv[2])*1000

title = ["Timestamp(ms)", "Timestamp(Str)", "Received Time(ms)", "Received Time(Str)",  "Delta Timestamp", "Rtime-Timestamp", "Timestamp Offset"]

data=[]
data.append(title)

stats={}
stats1={}
stats2={}
precision=int(sys.argv[3])
n=0
base=0
for line in file:
    tmpline=line.split()
    if n==0:
        base = int(tmpline[0])
        n+=1
        continue
       # tmpline.append(n)
    else:
        tmpline.append(int(tmpline[0])-(base+(n*sampleInterval)))
    tmpline[4] = int(tmpline[4])
    tmpline[5] = abs(int(tmpline[5]))
    data.append(tmpline)
    i = tmpline[4]//precision
    key = f"[{i*precision}-{(i+1)*precision})"
    if stats.get(key):
        stats[key] += 1
    else:
        stats[key] = 1
    i = abs(tmpline[4]-sampleInterval)//precision
    key = (i+1)*precision
    if stats1.get(key):
        stats1[key] += 1
    else:
        stats1[key] = 1
    i = abs(tmpline[6])//precision
    key = (i+1)*precision
    if stats2.get(key):
        stats2[key] += 1
    else:
        stats2[key] = 1
    
    n += 1

for row in data:
    ws.append(row)

ws1.append(['Relative Rang','Relative Number', 'Relative Percent'])

bardata=[]
for key in sorted(stats.keys()):
    tmpbarline=[]
    tmpbarline.append(key)
    tmpbarline.append(stats.get(key))
    tmpbarline.append(stats.get(key)*100/(n-1))
    bardata.append(tmpbarline)
    ws1.append(tmpbarline)
drawBarChart(ws1,2,len(bardata)+1,2,2,"F1")
drawPieChart(ws1,1,len(bardata)+1,3,3,"Q1",title="Relative sample interval (%)")

ws2.append(['Relative Rang', 'Number', 'Percent'])
piedata=[]
for key in sorted(stats1.keys()):
    tmppieline=[]
    tmppieline.append(key)
    tmppieline.append(stats1.get(key))
    tmppieline.append(stats1.get(key)*100/(n-1))
    piedata.append(tmppieline)
    ws2.append(tmppieline)
drawPieChart(ws2,1,len(piedata)+1,3,3,"E1")

ws3.append(['Absolute Rang', 'Number', 'Percent'])
piedata=[]
for key in sorted(stats2.keys()):
    tmppieline=[]
    tmppieline.append(key)
    tmppieline.append(stats2.get(key))
    tmppieline.append(stats2.get(key)*100/(n-1))
    piedata.append(tmppieline)
    ws3.append(tmppieline)
drawPieChart(ws3,1,len(piedata)+1,3,3,"E1")

drawLineChart(ws,1,len(data),5,5, "J1", True, title="Relative sample interval", ytitle="Interval (ms)",xtitle="Sample No.")
drawLineChart(ws,1,len(data),7,7, "J20",True, title="Sampling difference based on first update", ytitle="Delta interval (ms)",xtitle="Sample No.")

if len(sys.argv) > 4:
    i = 4
    while i<len(sys.argv):
        srow=sys.argv[i]
        if '-' in srow:
            selist=srow.split('-')
            st=int(selist[0])
            ed=int(selist[1])
        else:
            st=int(srow)
            ed=len(data)
        drawLineChart(ws,st,ed,5,6,"J"+str(20*(i-2)))
        i=i+1

wb.save(sys.argv[1]+".telemetry.xlsx")
file.close()
